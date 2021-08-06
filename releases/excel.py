import openpyxl
from .models import Release
from configuration.settings import MEDIA_ROOT
from django_countries.data import COUNTRIES

FORMATS = Release.Formats.values
STYLES = Release.BaseStyle.values

# There are 249 countries in dict, it may
COUNTRIES_DICT = {full_name: short_name for short_name, full_name in COUNTRIES.items()}


def save_excel_file(file, profile):
    """
    Parse, validate and save file. If there is error returns it, otherwise returns None
    """

    # you may say that file can be not excel format,
    # but i added file validation in ReleaseImportForm (forms.py)
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):

        # check if all cells in sheet not empty

        if not all([sheet.cell(row, col).value != "" for col in range(1, sheet.max_column + 1)]):
            return "There are empty values. Please check your excel file"

        try:
            # as django_countries saves country in DB with 2 chars ("New Zealand" -> "NZ")
            # and countries dict looks like {"NZ": "New Zealand"}
            valid_country = COUNTRIES_DICT[sheet.cell(row, 3).value]
        except:
            return f"Wrong country at {row} row, 3 column"

        try:
            # Convert DD.MM.YYYY format to YYYY-MM-DD
            valid_date = sheet.cell(row, 4).value.strftime("%Y-%m-%d")
        except:
            return f"wrong date format at {row} row, 4 column"

        format = sheet.cell(row, 6).value
        style = sheet.cell(row, 8).value

        try:
            release = Release(
                band_name=sheet.cell(row, 1).value,
                album_title=sheet.cell(row, 2).value,
                country=valid_country,
                release_date=valid_date,
                media_format_details=sheet.cell(row, 5).value,
                format=format if (format in FORMATS) else "Other",
                limited_edition=sheet.cell(row, 7).value,
                base_style=style if (style in STYLES) else "Other",
                profile=profile,
                sample=f"{MEDIA_ROOT}/audio/releases/dummy.mp3",
                cover_image=f"{MEDIA_ROOT}/images/releases/dummy.jpg",
                # Since a user can have several labels, for now I have done like this
                label=profile.label.first()

            )

            release.save()

        except:
            return "import failed. Your file may contain empty cells"